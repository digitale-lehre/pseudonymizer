#!/usr/bin/env python3
"""
Pseudonymisierung & De-Pseudonymisierung mit nur einem Secret.
Keine separate Key-Datei noetig. Unterstuetzt CSV und XLSX.

Verwendet AES-Verschluesselung (deterministisch): gleicher Secret + gleiche Daten
= gleiches Pseudonym. Nur wer den Secret kennt, kann zurueckfuehren.

Verwendung:
    # CSV pseudonymisieren
    python pseudonym.py encrypt eingabe.csv --secret "MeinGeheimnis"

    # XLSX pseudonymisieren
    python pseudonym.py encrypt eingabe.xlsx --secret "MeinGeheimnis"

    # De-Pseudonymisieren (selber Secret)
    python pseudonym.py decrypt eingabe_pseudo.xlsx --secret "MeinGeheimnis"

    # Mit Semikolon-CSV
    python pseudonym.py encrypt eingabe.csv --secret "MeinGeheimnis" --sep ";"
"""

import argparse
import base64
import csv
import hashlib
import hmac
import io
import re
import sys
from pathlib import Path

from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
from cryptography.hazmat.primitives import padding as sym_padding

__version__ = "0.4.0"

# Spaltennamen-Mapping: verschiedene Schreibweisen -> kanonischer Schluessel
# Jeder kanonische Schluessel hat eine Liste von moeglichen Spaltennamen
COLUMN_ALIASES = {
    "familienname": ["FAMILIENNAME", "Familienname", "familienname", "Zuname", "zuname",
                     "ZUNAME", "Nachname", "nachname", "NACHNAME",
                     "Last Name", "LastName", "Last name", "Lastname", "Surname", "surname",
                     "Family Name", "Family_Name",
                     "FAMILY_NAME_OF_STUDENT", "Family_Name_of_Student",
                     "Familienname oder Nachname",
                     "Familien- oder Nachname"],
    "vorname":      ["VORNAME", "Vorname", "vorname",
                     "First Name", "FirstName", "Firstname", "First name",
                     "Given Name", "GivenName", "Given name", "Rufname",
                     "FIRST_NAME_OF_STUDENT", "First_Name_of_Student"],
    "matnr":        ["MATRIKELNUMMER", "Matrikelnummer", "matrikelnummer", "Matnr", "matnr",
                     "MATNR", "MatrNr", "Matrikel", "matrikel", "MATRIKEL",
                     "Matrikelnr", "Matrikelnr.",
                     "StudentID", "Student_ID", "Student ID",
                     "REGISTRATION_NUMBER", "Registration_Number", "Registration Number",
                     "ID number", "ID Number", "ID-Nummer", "Kennnummer"],
    "email":        ["EMAIL_ADDRESS", "Email_Address", "E-Mail", "E-MAIL", "e-mail",
                     "Email", "email", "EMAIL", "Mail", "MAIL", "mail",
                     "E_MAIL", "EmailAddress", "email_address",
                     "E-Mail-Adresse", "E-Mail Adresse", "Emailadresse", "Mailadresse",
                     "Email address", "Email Address",
                     "E-Mail des Teilnehmers", "Attendee Email"],
    "pruefer":      ["Examiner", "EXAMINER", "examiner", "Prüfer", "PRÜFER", "prüfer",
                     "Pruefer", "PRUEFER", "pruefer", "Prufer", "PRUFER",
                     "Prüfer/in", "PrüferIn", "Prüfer:in"],
    "anzeigename":  ["Anzeigename", "ANZEIGENAME", "anzeigename",
                     "Display Name", "DisplayName", "DISPLAY NAME", "display name",
                     "Full Name", "FullName", "Student Name"],
    "svnr":         ["Sozialversicherungsnummer", "SOZIALVERSICHERUNGSNUMMER",
                     "SVNr", "SVNR", "SV-Nr", "SV-Nr.", "SV-Nummer", "SV Nummer",
                     "Versicherungsnummer", "Social Security Number", "SSN"],
    "geburtsdatum": ["Geburtsdatum", "GEBURTSDATUM", "Geburtstag", "GEBURTSTAG",
                     "Geb.Datum", "Geb.-Datum", "GebDatum", "Geb. Datum",
                     "Birthday", "Date of Birth", "DateOfBirth", "DOB",
                     "Birth Date", "BirthDate", "Birthdate"],
    "telefon":      ["Telefon", "TELEFON", "Telefonnummer", "TELEFONNUMMER",
                     "Tel", "Tel.", "Tel.Nr.", "TelNr",
                     "Phone", "Phone Number", "PhoneNumber",
                     "Handy", "Handynummer", "Mobilnummer", "Mobile",
                     "Mobiltelefon", "Cell", "Cell Phone"],
    # Valuatic Pruefungssoftware — eigene Keys, da examiner + candidate
    # in derselben Datei vorkommen und unabhaengig verschluesselt werden muessen
    "examiner_id":        ["examiner_id", "Examiner_ID", "EXAMINER_ID"],
    "examiner_nachname":  ["examiner_last_name", "Examiner_Last_Name", "EXAMINER_LAST_NAME"],
    "examiner_vorname":   ["examiner_first_name", "Examiner_First_Name", "EXAMINER_FIRST_NAME"],
    "candidate_id":       ["candidate_id", "Candidate_ID", "CANDIDATE_ID"],
    "candidate_nachname": ["candidate_last_name", "Candidate_Last_Name", "CANDIDATE_LAST_NAME"],
    "candidate_vorname":  ["candidate_first_name", "Candidate_First_Name", "CANDIDATE_FIRST_NAME"],
}
NAME_ALIASES = ["NAME", "Name", "name"]


def derive_key(secret: str) -> bytes:
    """Leitet einen 256-Bit AES-Key aus dem Secret ab (PBKDF2, 100k Runden)."""
    return hashlib.pbkdf2_hmac("sha256", secret.encode("utf-8"), b"MedUniWien-Pseudo-Salt", 100_000)


def deterministic_iv(key: bytes, plaintext: str) -> bytes:
    """Erzeugt einen deterministischen IV aus Key + Klartext (HMAC, 16 Bytes)."""
    return hmac.new(key, plaintext.encode("utf-8"), hashlib.sha256).digest()[:16]


def encrypt_value(key: bytes, plaintext: str) -> str:
    """Verschluesselt einen String -> URL-sicherer Base64-String."""
    if not plaintext:
        return plaintext
    iv = deterministic_iv(key, plaintext)
    padder = sym_padding.PKCS7(128).padder()
    padded = padder.update(plaintext.encode("utf-8")) + padder.finalize()
    cipher = Cipher(algorithms.AES(key), modes.CBC(iv))
    ct = cipher.encryptor().update(padded) + cipher.encryptor().finalize()
    raw = iv + ct
    return base64.urlsafe_b64encode(raw).rstrip(b"=").decode("ascii")


def decrypt_value(key: bytes, token: str) -> str:
    """Entschluesselt einen Base64-Token -> Klartext-String."""
    if not token:
        return token
    pad_len = 4 - len(token) % 4
    if pad_len < 4:
        token += "=" * pad_len
    try:
        raw = base64.urlsafe_b64decode(token.encode("ascii"))
    except Exception:
        return token
    if len(raw) < 32:
        return token
    iv = raw[:16]
    ct = raw[16:]
    try:
        cipher = Cipher(algorithms.AES(key), modes.CBC(iv))
        decryptor = cipher.decryptor()
        padded = decryptor.update(ct) + decryptor.finalize()
        unpadder = sym_padding.PKCS7(128).unpadder()
        plaintext = unpadder.update(padded) + unpadder.finalize()
        return plaintext.decode("utf-8")
    except Exception:
        return token


_SUFFIX_RE = re.compile(r'\.[a-zA-Z0-9_]+$')


def _normalize_header(h: str) -> str:
    """Entfernt Suffixe wie .x, .y, .1, .2 (z.B. aus R-Merges) von Spaltenheadern."""
    return _SUFFIX_RE.sub('', h)


def find_identity_cols(headers: list) -> dict:
    """Findet Identitaetsspalten anhand verschiedener Namenskonventionen (case-insensitive).
    Spaltenheader werden vor dem Matching normalisiert: Suffixe wie .x, .y (R-Merges)
    werden entfernt. Gibt dict zurueck: {kanonischer_key: tatsaechlicher_spaltenname}"""
    found = {}
    # Baue case-insensitive Lookup: lower(header) -> header
    # Versuche zuerst exakt, dann mit Suffix-Stripping
    header_lower = {h.strip().lower(): h for h in headers if h}
    header_norm = {_normalize_header(h.strip()).lower(): h for h in headers if h}
    for canon_key, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            al = alias.lower()
            if al in header_lower:
                found[canon_key] = header_lower[al]
                break
            if al in header_norm:
                found[canon_key] = header_norm[al]
                break
    return found


def find_name_col(headers: list) -> str:
    """Findet die NAME-Spalte (falls vorhanden, case-insensitive)."""
    header_lower = {h.strip().lower(): h for h in headers if h}
    for alias in NAME_ALIASES:
        if alias.lower() in header_lower:
            return header_lower[alias.lower()]
    return None


def find_header_row(rows, max_scan=20):
    """Findet die Header-Zeile in einer Liste von Zeilen (als Listen).
    Scannt bis zu max_scan Zeilen nach bekannten Identitaetsspalten.
    Gibt den Index zurueck (0-basiert). Fallback: 0 (erste Zeile)."""
    for i, row in enumerate(rows[:max_scan]):
        # row is a list of cell values (strings)
        headers = [str(c).strip() if c is not None else "" for c in row]
        if find_identity_cols(headers):
            return i
    return 0


# ======================== CSV ========================

def detect_file_encoding(raw: bytes) -> tuple:
    """Erkennt Encoding und BOM-Laenge aus rohen Bytes.
    Unterstuetzt UTF-8 (mit/ohne BOM), UTF-16 LE und UTF-16 BE.
    Gibt (encoding, bom_length) zurueck."""
    if len(raw) >= 2 and raw[:2] == b"\xff\xfe":
        return "utf-16-le", 2
    if len(raw) >= 2 and raw[:2] == b"\xfe\xff":
        return "utf-16-be", 2
    if len(raw) >= 3 and raw[:3] == b"\xef\xbb\xbf":
        return "utf-8", 3
    return "utf-8", 0


def process_csv(input_path: str, output_path: str, secret: str, mode: str, sep: str = ","):
    key = derive_key(secret)
    transform = encrypt_value if mode == "encrypt" else decrypt_value

    # Datei als Bytes lesen und Encoding erkennen (UTF-8, UTF-16 LE/BE)
    with open(input_path, "rb") as f:
        raw = f.read()

    encoding, bom_len = detect_file_encoding(raw)
    text = raw[bom_len:].decode(encoding)
    has_crlf = "\r\n" in text

    # Alle Zeilen als Listen parsen, um Header-Zeile zu finden
    reader_raw = csv.reader(io.StringIO(text), delimiter=sep)
    all_rows = list(reader_raw)

    if not all_rows:
        print("FEHLER: Datei ist leer.", file=sys.stderr)
        sys.exit(1)

    # Header-Zeile finden (kann nach Metadaten-Zeilen liegen, z.B. MLW-Export)
    header_idx = find_header_row(all_rows)
    fieldnames = all_rows[header_idx]
    data_rows_raw = all_rows[header_idx + 1:]

    line_terminator = "\r\n" if has_crlf else "\n"

    # Rohtext-Zeilen fuer Metadaten-Bewahrung und Quoting-Erkennung
    # (Einfache Zeilen ohne Newlines in Quotes — MLW-Metadaten sind sicher)
    raw_lines = text.split("\r\n") if has_crlf else text.split("\n")

    # Rohe Metadaten-Zeilen als Text bewahren (fuer byte-identische Ausgabe)
    meta_raw_text = ""
    if header_idx > 0:
        meta_raw_text = line_terminator.join(raw_lines[:header_idx]) + line_terminator

    # Quoting anhand der Header-Zeile erkennen (nicht erste Datei-Zeile!)
    header_line_raw = raw_lines[header_idx].strip() if header_idx < len(raw_lines) else ""
    quoting = csv.QUOTE_ALL if header_line_raw.startswith('"') else csv.QUOTE_MINIMAL

    # Konvertiere zu Dicts (wie DictReader)
    rows = []
    for raw_row in data_rows_raw:
        d = {}
        for i, h in enumerate(fieldnames):
            d[h] = raw_row[i] if i < len(raw_row) else ""
        rows.append(d)

    if not fieldnames:
        print("FEHLER: Konnte keine Spalten erkennen.", file=sys.stderr)
        sys.exit(1)

    id_cols = find_identity_cols(fieldnames)
    name_col = find_name_col(fieldnames)
    if not id_cols:
        print(f"FEHLER: Keine Identitaetsspalten gefunden.", file=sys.stderr)
        print(f"Vorhandene Spalten: {fieldnames}", file=sys.stderr)
        sys.exit(1)

    # NAME composite check (erkennt auch Reihenfolge: "Fam Vor" vs "Vor Fam")
    name_is_composite = False
    name_order = "fam_vor"
    fam_col = id_cols.get("familienname")
    vor_col = id_cols.get("vorname")
    if name_col and fam_col and vor_col and rows:
        sample = rows[0]
        name_val = (sample.get(name_col) or "").strip()
        fam_val = (sample.get(fam_col) or "").strip()
        vor_val = (sample.get(vor_col) or "").strip()
        if name_val == f"{fam_val} {vor_val}".strip():
            name_is_composite = True
            name_order = "fam_vor"
        elif name_val == f"{vor_val} {fam_val}".strip():
            name_is_composite = True
            name_order = "vor_fam"

    # Ergebnis in StringIO schreiben, dann mit Original-Encoding ausgeben
    str_out = io.StringIO()

    # Metadaten-Zeilen als Rohtext ausgeben (byte-identisch, nicht re-serialisiert)
    if meta_raw_text:
        str_out.write(meta_raw_text)

    # Header + Daten mit DictWriter
    dict_writer = csv.DictWriter(
        str_out, fieldnames=fieldnames, delimiter=sep,
        quoting=quoting, lineterminator=line_terminator
    )
    dict_writer.writeheader()
    for row in rows:
        new_row = dict(row)
        for canon_key, actual_col in id_cols.items():
            val = (row.get(actual_col) or "").strip()
            if val:
                new_row[actual_col] = transform(key, val)
        if name_col and name_is_composite:
            fam = new_row.get(fam_col, "")
            vor = new_row.get(vor_col, "")
            if name_order == "vor_fam":
                new_row[name_col] = f"{vor} {fam}".strip()
            else:
                new_row[name_col] = f"{fam} {vor}".strip()
        dict_writer.writerow(new_row)

    output_text = str_out.getvalue()
    with open(output_path, "wb") as f:
        if bom_len > 0:
            f.write(raw[:bom_len])
        f.write(output_text.encode(encoding))

    cols_info = ", ".join(f"{v}" for v in id_cols.values())
    if name_col and name_is_composite:
        cols_info += f" + {name_col}"
    action = "Pseudonymisierung" if mode == "encrypt" else "De-Pseudonymisierung"
    print(f"{action} abgeschlossen.")
    print(f"  Eingabe:    {input_path} ({len(rows)} Zeilen)")
    print(f"  Ausgabe:    {output_path}")
    print(f"  Spalten:    {cols_info}")
    print(f"  Encoding:   {encoding.upper()}{' (BOM)' if bom_len > 0 else ''}")
    if header_idx > 0:
        print(f"  Metadaten:  {header_idx} Zeile(n) vor Header (unveraendert)")
    print(f"  Verfahren:  AES-256-CBC (deterministisch, PBKDF2)")
    if mode == "encrypt":
        print(f"\n  Zum Zurueckfuehren:")
        print(f"  python pseudonym.py decrypt {output_path} --secret <IhrSecret>")


# ======================== XLSX ========================

def _fix_xlsx_drawings(input_path: str) -> str:
    """Repariert XLSX-Dateien mit fehlenden Drawing-Referenzen.
    Gibt Pfad zur (ggf. reparierten) Datei zurueck."""
    import zipfile
    import tempfile
    import os
    import re

    with zipfile.ZipFile(input_path, "r") as zf:
        names = zf.namelist()
        # Pruefe ob es Sheet-Rels gibt, die auf fehlende Drawings verweisen
        broken = False
        for name in names:
            if name.startswith("xl/worksheets/_rels/") and name.endswith(".rels"):
                content = zf.read(name).decode("utf-8")
                # Finde alle Drawing-Referenzen
                for m in re.finditer(r'Target="([^"]*drawing[^"]*)"', content):
                    target = m.group(1)
                    # Pfad relativ zu xl/worksheets/ aufloesen
                    if target.startswith("../"):
                        full_path = "xl/" + target[3:]
                    elif target.startswith("/"):
                        full_path = target.lstrip("/")
                    else:
                        full_path = "xl/worksheets/" + target
                    # Normalisiere Pfad
                    full_path = os.path.normpath(full_path).replace("\\", "/")
                    if full_path not in names:
                        broken = True
                        break
            if broken:
                break

        if not broken:
            return input_path

        # Reparatur: Drawing-Relationships aus Sheet-Rels entfernen
        print("  HINWEIS: Repariere fehlende Drawing-Referenzen in XLSX...")
        tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
        os.close(tmp_fd)

        with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zf.infolist():
                data = zf.read(item.filename)
                if item.filename.startswith("xl/worksheets/_rels/") and item.filename.endswith(".rels"):
                    content = data.decode("utf-8")
                    # Entferne Relationships mit Type=drawing, deren Target fehlt
                    cleaned = re.sub(
                        r'<Relationship[^>]*Type="[^"]*drawing[^"]*"[^>]*/>\s*',
                        '', content
                    )
                    data = cleaned.encode("utf-8")
                zout.writestr(item, data)

        return tmp_path


def process_xlsx(input_path: str, output_path: str, secret: str, mode: str):
    from openpyxl import load_workbook
    from copy import copy

    key = derive_key(secret)
    transform = encrypt_value if mode == "encrypt" else decrypt_value

    # Repariere fehlende Drawings falls noetig
    fixed_path = _fix_xlsx_drawings(input_path)
    try:
        wb = load_workbook(fixed_path)
    finally:
        if fixed_path != input_path:
            import os
            os.unlink(fixed_path)
    total_cells = 0
    sheets_processed = []
    sheets_skipped = []

    for ws in wb.worksheets:
        # Leere Sheets ueberspringen
        if ws.max_row is None or ws.max_row < 1:
            sheets_skipped.append(f"{ws.title} (leer)")
            continue

        # Alle Zeilen als Listen lesen, dann Header-Zeile finden
        all_rows = []
        for row_idx in range(1, (ws.max_row or 0) + 1):
            row_vals = []
            for col_idx in range(1, (ws.max_column or 0) + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                row_vals.append(str(val).strip() if val is not None else "")
            all_rows.append(row_vals)

        header_offset = find_header_row(all_rows)
        headers = all_rows[header_offset] if all_rows else []
        # header_row in openpyxl is 1-based: header_offset + 1
        header_row_num = header_offset + 1

        id_cols = find_identity_cols(headers)
        if not id_cols:
            # Warnung ausgeben, aber nicht abbrechen
            sheets_skipped.append(f"{ws.title} (keine Identitaetsspalten erkannt: {[h for h in headers if h]})")
            continue

        data_start_row = header_row_num + 1
        if ws.max_row < data_start_row:
            sheets_skipped.append(f"{ws.title} (nur Header, keine Daten)")
            continue

        name_col_name = find_name_col(headers)

        # Spaltenindizes ermitteln (1-basiert fuer openpyxl)
        col_indices = {}  # canon_key -> col_index (1-basiert)
        for canon_key, actual_name in id_cols.items():
            col_indices[canon_key] = headers.index(actual_name) + 1

        name_col_idx = None
        if name_col_name and name_col_name in headers:
            name_col_idx = headers.index(name_col_name) + 1

        # NAME composite check (anhand erster Datenzeile, erkennt Reihenfolge)
        name_is_composite = False
        name_order = "fam_vor"
        fam_idx = col_indices.get("familienname")
        vor_idx = col_indices.get("vorname")
        if name_col_idx and fam_idx and vor_idx and ws.max_row >= data_start_row:
            name_val = str(ws.cell(row=data_start_row, column=name_col_idx).value or "").strip()
            fam_val = str(ws.cell(row=data_start_row, column=fam_idx).value or "").strip()
            vor_val = str(ws.cell(row=data_start_row, column=vor_idx).value or "").strip()
            if name_val == f"{fam_val} {vor_val}".strip():
                name_is_composite = True
                name_order = "fam_vor"
            elif name_val == f"{vor_val} {fam_val}".strip():
                name_is_composite = True
                name_order = "vor_fam"

        sheet_count = 0
        for row_idx in range(data_start_row, ws.max_row + 1):
            for canon_key, ci in col_indices.items():
                cell = ws.cell(row=row_idx, column=ci)
                val = cell.value
                if val is not None:
                    val_str = str(val).strip()
                    if val_str:
                        cell.value = transform(key, val_str)
                        sheet_count += 1

            # NAME-Spalte aktualisieren (Original-Reihenfolge beibehalten)
            if name_col_idx and name_is_composite and fam_idx and vor_idx:
                fam_new = str(ws.cell(row=row_idx, column=fam_idx).value or "").strip()
                vor_new = str(ws.cell(row=row_idx, column=vor_idx).value or "").strip()
                if name_order == "vor_fam":
                    ws.cell(row=row_idx, column=name_col_idx).value = f"{vor_new} {fam_new}".strip()
                else:
                    ws.cell(row=row_idx, column=name_col_idx).value = f"{fam_new} {vor_new}".strip()

        total_cells += sheet_count
        cols_info = ", ".join(id_cols.values())
        if name_is_composite and name_col_name:
            cols_info += f" + {name_col_name}"
        sheets_processed.append(f"{ws.title} ({cols_info}: {sheet_count} Zellen)")
        if header_row_num > 1:
            print(f"    (Header in Zeile {header_row_num}, {header_row_num - 1} Metadaten-Zeile(n) uebersprungen)")

    wb.save(output_path)

    action = "Pseudonymisierung" if mode == "encrypt" else "De-Pseudonymisierung"
    print(f"{action} abgeschlossen.")
    print(f"  Eingabe:    {input_path}")
    print(f"  Ausgabe:    {output_path}")
    print(f"  Verfahren:  AES-256-CBC (deterministisch, PBKDF2)")
    print(f"  Sheets:     {len(sheets_processed)} verarbeitet, {total_cells} Zellen")
    for s in sheets_processed:
        print(f"    + {s}")
    if sheets_skipped:
        print(f"  Uebersprungen: {len(sheets_skipped)}")
        for s in sheets_skipped:
            print(f"    - {s}")
    if not sheets_processed:
        print("  WARNUNG: Kein Sheet mit erkennbaren Identitaetsspalten gefunden!")
        print("  Erkannte Spaltennamen: FAMILIENNAME/Familienname/Zuname/Nachname,")
        print("    VORNAME/Vorname, MATRIKELNUMMER/Matrikelnummer/Matnr (case-insensitive)")
    if mode == "encrypt":
        print(f"\n  Zum Zurueckfuehren:")
        print(f"  python pseudonym.py decrypt {output_path} --secret <IhrSecret>")


# ======================== BATCH HELPERS ========================

def collect_input_files(paths: list) -> list:
    """Sammelt Eingabedateien. ZIP-Archive werden entpackt (nur CSV/XLSX).
    Gibt Liste von Path-Objekten zurueck (ggf. in tempdir extrahiert)."""
    import zipfile
    import tempfile

    SUPPORTED = {".csv", ".tsv", ".txt", ".xlsx"}
    ZIP_EXTRACT = {".csv", ".tsv", ".xlsx"}
    collected = []

    for p in paths:
        p = Path(p)
        if p.suffix.lower() == ".zip":
            tmpdir = Path(tempfile.mkdtemp(prefix="pseudo_zip_"))
            with zipfile.ZipFile(p, "r") as zf:
                for member in zf.namelist():
                    if member.endswith("/") or Path(member).name.startswith("."):
                        continue
                    ext = Path(member).suffix.lower()
                    if ext in ZIP_EXTRACT:
                        target = tmpdir / Path(member).name
                        counter = 1
                        while target.exists():
                            stem = Path(member).stem
                            target = tmpdir / f"{stem}_{counter}{ext}"
                            counter += 1
                        with open(target, "wb") as f:
                            f.write(zf.read(member))
                        collected.append(target)
            if not any(f.parent == tmpdir for f in collected):
                import shutil
                shutil.rmtree(tmpdir, ignore_errors=True)
        elif p.suffix.lower() in SUPPORTED:
            collected.append(p)

    return collected


def create_output_zip(results: list, zip_path: str):
    """Buendelt Ergebnisdateien in ein ZIP-Archiv.
    results: Liste von (input_path, output_path) Tupeln."""
    import zipfile

    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for _, output_path in results:
            zf.write(output_path, Path(output_path).name)
    print(f"\n  ZIP erstellt: {zip_path}")


def make_output_path(input_path, mode: str, output_dir: str = None) -> str:
    """Erzeugt Ausgabepfad: <name>_pseudo.<ext> bzw. <name>_restored.<ext>.
    Optional in ein anderes Verzeichnis (output_dir)."""
    p = Path(input_path)
    suffix = "_pseudo" if mode == "encrypt" else "_restored"
    name = f"{p.stem}{suffix}{p.suffix}"
    if output_dir:
        return str(Path(output_dir) / name)
    return str(p.with_name(name))


# ======================== DISPATCH ========================

def process_file(input_path: str, output_path: str, secret: str, mode: str, sep: str = ","):
    """Verarbeitet eine einzelne CSV/XLSX-Datei (Dispatch nach Dateityp)."""
    ext = Path(input_path).suffix.lower()
    if ext == ".xlsx":
        process_xlsx(input_path, output_path, secret, mode)
    elif ext in (".csv", ".tsv", ".txt"):
        process_csv(input_path, output_path, secret, mode, sep)
    else:
        raise ValueError(f"Unbekanntes Dateiformat '{ext}'. Unterstuetzt: .csv, .tsv, .txt, .xlsx")


# ======================== MAIN ========================

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="CSV/XLSX pseudonymisieren/de-pseudonymisieren — nur mit Secret, ohne Key-Datei"
    )
    parser.add_argument("--version", action="version", version=f"%(prog)s {__version__}")
    parser.add_argument("mode", choices=["encrypt", "decrypt"],
                        help="encrypt = pseudonymisieren, decrypt = zurueckfuehren")
    parser.add_argument("input", nargs="+",
                        help="Pfad(e) zu CSV/XLSX/ZIP-Datei(en)")
    parser.add_argument("--secret", required=True, help="Ihr geheimer Schluessel")
    parser.add_argument("--output", "-o",
                        help="Ausgabepfad (nur bei einzelner Datei)")
    parser.add_argument("--output-dir",
                        help="Ausgabeverzeichnis fuer Batch-Modus")
    parser.add_argument("--zip", action="store_true",
                        help="Ergebnis als ZIP-Datei buendeln")
    parser.add_argument("--sep", "-s", default=",",
                        help="CSV-Trennzeichen (Standard: Komma, wird bei XLSX ignoriert)")
    args = parser.parse_args()

    # --output nur bei einzelner Datei
    if args.output and len(args.input) > 1:
        print("FEHLER: --output nur bei einzelner Datei moeglich. "
              "Verwenden Sie --output-dir fuer Batch.", file=sys.stderr)
        sys.exit(1)

    # Pruefen ob alle Eingabedateien existieren
    for p in args.input:
        if not Path(p).exists():
            print(f"FEHLER: Datei nicht gefunden: {p}", file=sys.stderr)
            sys.exit(1)

    # Dateien sammeln (ZIP-Archive werden entpackt)
    all_files = collect_input_files(args.input)
    if not all_files:
        print("FEHLER: Keine verarbeitbaren Dateien gefunden (CSV/XLSX).", file=sys.stderr)
        sys.exit(1)

    if len(all_files) > 1:
        print(f"\nBatch-Modus: {len(all_files)} Datei(en)\n")

    results = []
    for f in all_files:
        if args.output and len(all_files) == 1:
            out = args.output
        else:
            out = make_output_path(f, args.mode, args.output_dir)
        try:
            process_file(str(f), out, args.secret, args.mode, args.sep)
            results.append((str(f), out, True, None))
        except Exception as e:
            print(f"\nFEHLER bei {f.name}: {e}", file=sys.stderr)
            results.append((str(f), None, False, str(e)))

    # ZIP-Ausgabe
    if args.zip:
        successful = [(r[0], r[1]) for r in results if r[2]]
        if successful:
            first_input = Path(successful[0][0])
            if args.output_dir:
                zip_dir = args.output_dir
            else:
                zip_dir = str(first_input.parent)
            zip_name = str(Path(zip_dir) / f"batch_{args.mode}_{len(successful)}files.zip")
            create_output_zip(successful, zip_name)

    # Zusammenfassung bei Batch
    if len(all_files) > 1:
        ok = sum(1 for r in results if r[2])
        fail = len(results) - ok
        print(f"\n{'='*50}")
        print(f"Batch abgeschlossen: {ok} erfolgreich, {fail} fehlgeschlagen")
        if fail > 0:
            for inp, _, success, err in results:
                if not success:
                    print(f"  FEHLER: {Path(inp).name}: {err}")
        print(f"{'='*50}")

    # Temp-Verzeichnisse aufraeumen (aus ZIP-Extraktion)
    import shutil
    cleaned = set()
    for f in all_files:
        if "pseudo_zip_" in str(f.parent):
            d = str(f.parent)
            if d not in cleaned:
                shutil.rmtree(d, ignore_errors=True)
                cleaned.add(d)

    if any(not r[2] for r in results):
        sys.exit(1)
