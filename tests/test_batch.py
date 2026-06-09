import sys
import zipfile
import io
from pathlib import Path

from openpyxl import Workbook, load_workbook

sys.path.insert(0, str(Path(__file__).parent.parent))
from pseudonym import process_file, make_output_path, collect_input_files, create_output_zip


def _build_xlsx_bytes(rows):
    """rows: list of lists (first row = header). Returns .xlsx bytes."""
    wb = Workbook()
    ws = wb.active
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_xlsm(path, rows):
    """Create an .xlsm: a normal workbook + an injected dummy xl/vbaProject.bin.
    openpyxl 3.1.5 ignores the unreferenced part on load and re-merges it on
    save when keep_vba=True."""
    xlsx_bytes = _build_xlsx_bytes(rows)
    with zipfile.ZipFile(io.BytesIO(xlsx_bytes), "r") as src, \
         zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as out:
        for item in src.infolist():
            out.writestr(item, src.read(item.filename))
        out.writestr("xl/vbaProject.bin", b"DUMMY-VBA")


def test_process_file_csv(tmp_path):
    """process_file() dispatches CSV correctly and produces encrypted output."""
    src = tmp_path / "test.csv"
    src.write_text("Vorname,Familienname\nMax,Mustermann\n", encoding="utf-8")
    dst = tmp_path / "test_pseudo.csv"
    process_file(str(src), str(dst), "geheim", "encrypt", ",")
    assert dst.exists()
    content = dst.read_text(encoding="utf-8")
    assert "Max" not in content
    assert "Mustermann" not in content
    assert "Vorname" in content
    assert "Familienname" in content


def test_process_file_roundtrip(tmp_path):
    """Encrypt then decrypt produces original content."""
    original = "Vorname,Familienname\nMax,Mustermann\nEva,Testerin\n"
    src = tmp_path / "data.csv"
    src.write_text(original, encoding="utf-8")
    enc = tmp_path / "data_pseudo.csv"
    process_file(str(src), str(enc), "secret123", "encrypt", ",")
    dec = tmp_path / "data_restored.csv"
    process_file(str(enc), str(dec), "secret123", "decrypt", ",")
    restored = dec.read_text(encoding="utf-8")
    assert "Max" in restored
    assert "Mustermann" in restored
    assert "Eva" in restored
    assert "Testerin" in restored


def test_make_output_path_encrypt():
    assert make_output_path(Path("/d/students.csv"), "encrypt") == "/d/students_pseudo.csv"


def test_make_output_path_decrypt():
    assert make_output_path(Path("/d/data_pseudo.csv"), "decrypt") == "/d/data_pseudo_restored.csv"


def test_make_output_path_xlsx():
    assert make_output_path(Path("/d/data.xlsx"), "encrypt") == "/d/data_pseudo.xlsx"


def test_make_output_path_with_output_dir(tmp_path):
    result = make_output_path(Path("/d/students.csv"), "encrypt", str(tmp_path))
    assert result == str(tmp_path / "students_pseudo.csv")


def test_collect_input_files_plain(tmp_path):
    f1 = tmp_path / "a.csv"
    f1.write_text("Vorname\nMax\n")
    f2 = tmp_path / "b.csv"
    f2.write_text("Vorname\nEva\n")
    result = collect_input_files([str(f1), str(f2)])
    assert [Path(r) for r in result] == [f1, f2]


def test_collect_input_files_zip(tmp_path):
    csv1 = tmp_path / "data.csv"
    csv1.write_text("Vorname\nMax\n")
    txt = tmp_path / "readme.txt"
    txt.write_text("ignore me")
    zp = tmp_path / "bundle.zip"
    with zipfile.ZipFile(zp, "w") as zf:
        zf.write(csv1, "data.csv")
        zf.write(txt, "readme.txt")
    result = collect_input_files([str(zp)])
    assert len(result) == 1
    assert result[0].name == "data.csv"


def test_collect_input_files_mixed(tmp_path):
    plain = tmp_path / "plain.csv"
    plain.write_text("Vorname\nEva\n")
    inner = tmp_path / "inner.csv"
    inner.write_text("Vorname\nMax\n")
    zp = tmp_path / "archive.zip"
    with zipfile.ZipFile(zp, "w") as zf:
        zf.write(inner, "inner.csv")
    result = collect_input_files([str(plain), str(zp)])
    assert len(result) == 2
    names = [r.name for r in result]
    assert "plain.csv" in names
    assert "inner.csv" in names


def test_create_output_zip(tmp_path):
    f1 = tmp_path / "a_pseudo.csv"
    f1.write_text("encrypted_a")
    f2 = tmp_path / "b_pseudo.csv"
    f2.write_text("encrypted_b")
    zip_path = tmp_path / "output.zip"
    create_output_zip([(str(f1), str(f1)), (str(f2), str(f2))], str(zip_path))
    assert zip_path.exists()
    with zipfile.ZipFile(zip_path, "r") as zf:
        names = zf.namelist()
        assert "a_pseudo.csv" in names
        assert "b_pseudo.csv" in names
        assert zf.read("a_pseudo.csv").decode() == "encrypted_a"


def test_extra_cols_encrypts_custom_column(tmp_path):
    """Extra columns are encrypted alongside auto-detected ones."""
    src = tmp_path / "data.csv"
    src.write_text("Vorname,Familienname,Kommentar\nMax,Muster,Geheim\n", encoding="utf-8")
    dst = tmp_path / "data_pseudo.csv"
    process_file(str(src), str(dst), "secret", "encrypt", ",", extra_cols=["Kommentar"])
    content = dst.read_text(encoding="utf-8")
    assert "Max" not in content
    assert "Muster" not in content
    assert "Geheim" not in content
    assert "Vorname" in content
    assert "Kommentar" in content


def test_extra_cols_roundtrip(tmp_path):
    """Extra-col encrypt then decrypt restores original."""
    src = tmp_path / "data.csv"
    src.write_text("Vorname,Familienname,Notiz\nEva,Test,Vertraulich\n", encoding="utf-8")
    enc = tmp_path / "data_pseudo.csv"
    process_file(str(src), str(enc), "s", "encrypt", ",", extra_cols=["Notiz"])
    dec = tmp_path / "data_restored.csv"
    process_file(str(enc), str(dec), "s", "decrypt", ",", extra_cols=["Notiz"])
    content = dec.read_text(encoding="utf-8")
    assert "Vertraulich" in content


def test_extra_cols_missing_column_ignored(tmp_path):
    """Extra column that doesn't exist in file is silently ignored."""
    src = tmp_path / "data.csv"
    src.write_text("Vorname,Familienname\nMax,Muster\n", encoding="utf-8")
    dst = tmp_path / "data_pseudo.csv"
    process_file(str(src), str(dst), "secret", "encrypt", ",", extra_cols=["NichtVorhanden"])
    assert dst.exists()
    content = dst.read_text(encoding="utf-8")
    assert "Max" not in content


def test_extra_cols_no_duplicates(tmp_path):
    """If extra col is already auto-detected, don't encrypt twice."""
    src = tmp_path / "data.csv"
    src.write_text("Vorname,Familienname\nMax,Muster\n", encoding="utf-8")
    dst = tmp_path / "data_pseudo.csv"
    process_file(str(src), str(dst), "secret", "encrypt", ",", extra_cols=["Vorname"])
    assert dst.exists()


def test_xlsx_name_column_standalone_encrypted(tmp_path):
    """XLSX: standalone 'Name' column (no Familienname) is encrypted + round-trips."""
    src = tmp_path / "data.xlsx"
    src.write_bytes(_build_xlsx_bytes([["Name", "Vorname"], ["Mustermann", "Max"]]))
    enc = tmp_path / "data_pseudo.xlsx"
    process_file(str(src), str(enc), "secret", "encrypt")
    ws = load_workbook(enc).active
    assert ws.cell(row=2, column=1).value != "Mustermann"  # Name encrypted
    assert ws.cell(row=2, column=2).value != "Max"          # Vorname encrypted
    dec = tmp_path / "data_restored.xlsx"
    process_file(str(enc), str(dec), "secret", "decrypt")
    ws2 = load_workbook(dec).active
    assert ws2.cell(row=2, column=1).value == "Mustermann"
    assert ws2.cell(row=2, column=2).value == "Max"


def test_xlsx_name_column_only_sheet(tmp_path):
    """XLSX: a sheet whose only recognized column is 'Name' is still processed."""
    src = tmp_path / "names.xlsx"
    src.write_bytes(_build_xlsx_bytes([["Name"], ["Mustermann"], ["Testerin"]]))
    enc = tmp_path / "names_pseudo.xlsx"
    process_file(str(src), str(enc), "secret", "encrypt")
    ws = load_workbook(enc).active
    assert ws.cell(row=2, column=1).value != "Mustermann"  # encrypted
    assert ws.cell(row=3, column=1).value != "Testerin"
    dec = tmp_path / "names_restored.xlsx"
    process_file(str(enc), str(dec), "secret", "decrypt")
    ws2 = load_workbook(dec).active
    assert ws2.cell(row=2, column=1).value == "Mustermann"  # restored
    assert ws2.cell(row=3, column=1).value == "Testerin"


def test_name_column_standalone_encrypted(tmp_path):
    """A 'Name' column without separate Vorname/Familienname is encrypted whole."""
    src = tmp_path / "data.csv"
    src.write_text("Name,Vorname\nMustermann,Max\n", encoding="utf-8")
    dst = tmp_path / "data_pseudo.csv"
    process_file(str(src), str(dst), "secret", "encrypt", ",")
    content = dst.read_text(encoding="utf-8")
    assert "Mustermann" not in content
    assert "Max" not in content
    assert "Name" in content and "Vorname" in content


def test_name_column_standalone_roundtrip(tmp_path):
    """Standalone 'Name' column round-trips exactly."""
    src = tmp_path / "data.csv"
    src.write_text("Name,Vorname\nMustermann,Max\nTesterin,Eva\n", encoding="utf-8")
    enc = tmp_path / "data_pseudo.csv"
    process_file(str(src), str(enc), "secret123", "encrypt", ",")
    dec = tmp_path / "data_restored.csv"
    process_file(str(enc), str(dec), "secret123", "decrypt", ",")
    restored = dec.read_text(encoding="utf-8")
    assert "Mustermann" in restored and "Max" in restored
    assert "Testerin" in restored and "Eva" in restored


def test_name_not_composite_fallback_encrypted(tmp_path):
    """Name present alongside Vorname+Familienname but composite check fails
    (title) -> Name is still encrypted as a whole value, round-trips exactly."""
    src = tmp_path / "data.csv"
    src.write_text(
        "Vorname,Familienname,Name\nMax,Mustermann,Dr. Max Mustermann\n",
        encoding="utf-8",
    )
    enc = tmp_path / "data_pseudo.csv"
    process_file(str(src), str(enc), "secret", "encrypt", ",")
    content = enc.read_text(encoding="utf-8")
    assert "Dr. Max Mustermann" not in content
    assert "Mustermann" not in content and "Max" not in content
    dec = tmp_path / "data_restored.csv"
    process_file(str(enc), str(dec), "secret", "decrypt", ",")
    assert "Dr. Max Mustermann" in dec.read_text(encoding="utf-8")


def test_name_composite_still_recomposed(tmp_path):
    """Regression: a real composite Name is recomposed from encrypted parts
    (value contains a space), not encrypted as one token."""
    src = tmp_path / "data.csv"
    src.write_text(
        "Vorname,Familienname,Name\nMax,Mustermann,Mustermann Max\n",
        encoding="utf-8",
    )
    enc = tmp_path / "data_pseudo.csv"
    process_file(str(src), str(enc), "secret", "encrypt", ",")
    name_field = enc.read_text(encoding="utf-8").splitlines()[1].split(",")[2]
    assert " " in name_field
    dec = tmp_path / "data_restored.csv"
    process_file(str(enc), str(dec), "secret", "decrypt", ",")
    assert "Mustermann Max" in dec.read_text(encoding="utf-8")


def test_name_column_only_no_other_identity(tmp_path):
    """A CSV whose only recognized column is 'Name' is still encrypted (not rejected)."""
    src = tmp_path / "names.csv"
    src.write_text("Name\nMustermann\nTesterin\n", encoding="utf-8")
    enc = tmp_path / "names_pseudo.csv"
    process_file(str(src), str(enc), "secret", "encrypt", ",")
    content = enc.read_text(encoding="utf-8")
    assert "Mustermann" not in content
    assert "Testerin" not in content
    assert "Name" in content
    dec = tmp_path / "names_restored.csv"
    process_file(str(enc), str(dec), "secret", "decrypt", ",")
    restored = dec.read_text(encoding="utf-8")
    assert "Mustermann" in restored and "Testerin" in restored


def test_xlsx_name_not_composite_fallback(tmp_path):
    """XLSX: Name alongside Vorname+Familienname but composite check fails
    (title) -> Name still encrypted as a whole single-token value, round-trips."""
    src = tmp_path / "data.xlsx"
    src.write_bytes(_build_xlsx_bytes(
        [["Vorname", "Familienname", "Name"],
         ["Max", "Mustermann", "Dr. Max Mustermann"]]))
    enc = tmp_path / "data_pseudo.xlsx"
    process_file(str(src), str(enc), "secret", "encrypt")
    ws = load_workbook(enc).active
    name_val = ws.cell(row=2, column=3).value
    assert name_val != "Dr. Max Mustermann"   # Name encrypted
    assert " " not in name_val                 # single token, NOT recomposed
    dec = tmp_path / "data_restored.xlsx"
    process_file(str(enc), str(dec), "secret", "decrypt")
    ws2 = load_workbook(dec).active
    assert ws2.cell(row=2, column=3).value == "Dr. Max Mustermann"
